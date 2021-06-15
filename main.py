import sys
import os
import time
import traceback
from io import BytesIO
from pathlib import Path

import requests
from PIL import Image
from bs4 import BeautifulSoup as BS
from fake_useragent import UserAgent
from openpyxl import load_workbook

# ============ НАСТРОЙКИ ============
INPUT_FILENAME = 'barcodes.xlsx'
STATUS_MAP = {
    'available': 'В наличии',
    'unavailable': 'Нет в наличии',
    'out_of_stock': 'Нет в наличии'
}
REQUESTS_TIMEOUT = 0.2
DOWNLOAD_IMAGES = True
IMAGES_DIRECTORY = 'images'
# ============ КОНЕЦ НАСТРОЕК ============

IMAGES_PATH = Path(IMAGES_DIRECTORY)
IMAGES_PATH.mkdir(parents=True, exist_ok=True)


def get_workbook(filename):
    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        print('[!] Входной файл `{}` не найден.'.format(filename))
        sys.exit(1)
    else:
        return workbook


def get_search_results(text):
    """Сделать GET запрос на эндпоинт поиска и получить результат в виде словаря python."""
    url = f'https://search.rozetka.com.ua/search/api/v6/?front-type=xl&country=UA&lang=ru&text={text}'
    headers = {
        'user-agent': UserAgent().random
    }

    response = requests.get(url, headers=headers)
    return response.json()


def find_product(barcode, article):
    def get_product(products):
        """Найти товар в названии которого присутствует заданный артикул и вернуть его."""
        for prod in products.get('data', {}).get('goods', []):
            if article in prod['title']:
                return prod

    search_results = get_search_results(barcode)
    product = get_product(search_results)
    # Если нашли товар, то вернуть его.
    if product is not None:
        return product

    # Если товар по штрихкоду не найден, то поиск по артикулу.
    search_results = get_search_results(article)
    product = get_product(search_results)
    if product is not None:
        return product


def get_images_from_product(url, barcode):
    def download_image(image_url, filename):
        """Скачать и сохранить изображение по заданной ссылке и вернуть путь к файлу."""
        try:
            response = requests.get(image_url, stream=True)

            if not response.ok:
                return

            path = IMAGES_PATH.joinpath(filename)
            # Сохранить последовательность байтов в виде изображения на компьютер.
            Image.open(BytesIO(response.content)).save(path)
        except Exception as err:
            print('[!] Ошибка в `download_image()`:')
            traceback.print_tb(err.__traceback__)
        else:
            return path

    try:
        content = requests.get(url).content
        # интерфейс доступа к html.
        soup = BS(content, 'html.parser')

        filenames = []
        for num, image in enumerate(soup.find_all('img', class_='product-thumbnails__picture'), start=1):
            url = image['src']
            # Заменить preview на big, чтобы получить фото высокогго разрешения
            url_bits = url.split('/')
            url_bits[-2] = 'big'
            url = '/'.join(url_bits)

            # Получть расширение файла
            ext = os.path.splitext(url)[1]

            output = '{barcode}_{num}{ext}'.format(barcode=barcode, num=num, ext=ext)
            image_path = download_image(url, output)
            if image_path:
                filenames.append(str(image_path))

    except Exception as err:
        print('[!] Ошибка в `download_image()`:')
        traceback.print_tb(err.__traceback__)
    else:
        return filenames


def main():
    workbook = get_workbook(INPUT_FILENAME)
    sheet = workbook.active

    try:
        for row in sheet.iter_rows(min_row=2, max_col=4, max_row=sheet.max_row):
            barcode = row[0].value
            article = row[1].value

            if not (barcode and article):
                continue

            try:
                product = find_product(barcode, article)
            except Exception as err:
                print('[!] Ошибка в `find_product()` (ШК {}):'.format(barcode))
                traceback.print_tb(err.__traceback__)
                continue

            if product:
                # Получить "красивое" значение статуса продукта.
                # Если статус не найден, вернуть его как есть.
                status = STATUS_MAP.get(product['status'], product['status'])
            else:
                status = 'Не найден'

            row[2].value = status

            if DOWNLOAD_IMAGES and product and 'href' in product:
                filenames = get_images_from_product(product['href'], barcode)
                if filenames:
                    row[3].value = '\n'.join(filenames)

            print('[+]', barcode, status)
            time.sleep(REQUESTS_TIMEOUT)
    except Exception as err:
        print('[!] Ошибка в `main()`:')
        traceback.print_tb(err.__traceback__)
    finally:
        workbook.save(INPUT_FILENAME)


if __name__ == '__main__':
    main()
